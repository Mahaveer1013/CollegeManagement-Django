from openpyxl.styles import PatternFill
import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
import json
import requests
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.shortcuts import (HttpResponse, get_object_or_404, redirect, render)
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import UpdateView
import csv
from django.views.decorators.http import require_GET
from .forms import *
from .models import *
import datetime
from django.utils import timezone
from django.db.models import Count


def admin_home(request):
    total_staff = Staff.objects.all().count()
    total_students = Student.objects.all().count()
    subjects = Subject.objects.all()
    total_subject = subjects.count()
    total_department = Department.objects.all().count()
    attendance_list = Attendance.objects.filter(subject__in=subjects)
    total_attendance = attendance_list.count()
    attendance_list = []
    subject_list = []
    for subject in subjects:
        attendance_count = Attendance.objects.filter(subject=subject).count()
        subject_list.append(subject.name[:7])
        attendance_list.append(attendance_count)

    # Total Subjects and students in Each Department
    department_all = Department.objects.all()
    department_name_list = []
    # subject_count_list = []
    student_count_list_in_department = []

    for department in department_all:
        # subjects = Subject.objects.filter(department_id=department.id).count()
        students = Student.objects.filter(department_id=department.id).count()
        department_name_list.append(department.name)
        # subject_count_list.append(subjects)
        student_count_list_in_department.append(students)

    subject_all = Subject.objects.all()
    subject_list = []
    student_count_list_in_subject = Subject.objects.count()
    for subject in subject_all:
        subject_list.append(subject.name)

    # For Students
    student_attendance_present_list = []
    student_attendance_leave_list = []
    student_name_list = []

    students = Student.objects.all()
    for student in students:

        attendance = AttendanceReport.objects.filter(
            student_id=student.id, status=True).count()
        absent = AttendanceReport.objects.filter(
            student_id=student.id, status=False).count()
        leave = LeaveReportStudent.objects.filter(
            student_id=student.id, status=1).count()
        student_attendance_present_list.append(attendance)
        student_attendance_leave_list.append(leave+absent)
        student_name_list.append(student.user.first_name)

    context = {
        'page_title': "Administrative Dashboard",
        'total_students': total_students,
        'total_staff': total_staff,
        'total_department': total_department,
        'total_subject': total_subject,
        'subject_list': subject_list,
        'attendance_list': attendance_list,
        'student_attendance_present_list': [10,8,7],
        'student_attendance_leave_list': [10,8,7],
        "student_name_list": ['mahaveer','fvrebrt','egtbv'],
        "student_count_list_in_subject": student_count_list_in_subject,
        "student_count_list_in_department": student_count_list_in_department,
        "department_name_list": department_name_list,

    }
    return render(request, 'hod_template/home_content.html', context)


def get_classes_by_department(request):
    department_id = request.GET.get('department_id')
    classes = ClassList.objects.filter(
        department_id=department_id).values('id', 'semester', 'section')
    return JsonResponse(list(classes), safe=False)


def get_classes_by_exam(request):
    exam_type = request.GET.get('exam_type')
    exam_detail = ExamDetail.objects.filter(id=exam_type).first()
    classes = ClassList.objects.filter(department=exam_detail.department, semester=exam_detail.semester).values('id','department__name' ,'semester', 'section')
    return JsonResponse(list(classes), safe=False)


def admin_view_attendance(request):
    user = get_object_or_404(Admin, user=request.user)
    form = AttendanceSelectionForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            date = form.cleaned_data.get('date')
            class_id = form.cleaned_data.get('class_name')
            students_id = Student.objects.filter(class_name=class_id)
            attendance_array = []
            if len(students_id) > 0:
                for student in students_id:
                    attendance = Attendance.objects.filter(
                        date=date, student=student)
                    if attendance:
                        one_student_attendance = []
                        one_student_attendance.append(
                            attendance[0].student.roll_number)
                        for period in range(1, 9):
                            student_attendance_status = Attendance.objects.filter(
                                student=student, date=date, period=period).first()
                            if student_attendance_status:
                                one_student_attendance.append(
                                    student_attendance_status.status)
                            else:
                                one_student_attendance.append(4)
                        attendance_array.append(one_student_attendance) if len(
                            one_student_attendance) > 0 else None
            else:
                messages.success(request, "No Attendance Records Found")
                return redirect(reverse('admin_view_attendance'))

            return render(request, "hod_template/attendance_view_page.html", {'attendance_list': attendance_array, 'class_name': str(class_id)})
        else:
            messages.error(request, "Invalid Data Provided")
    context = {'form': form, 'page_title': 'Select Attendance Details'}
    return render(request, "hod_template/admin_view_attendance.html", context)


def get_staff_profile(request):
    if request.method!='POST':
        return render(request, 'hod_template/get_profile.html',{'type':'staff'})
    else:
        staff_id = request.POST.get('faculty_id')
        print(staff_id)
        staff = Staff.objects.filter(faculty_id = staff_id).first()
        if staff:
            print(staff)
            return render(request, 'staff_template/staff_profile.html',{'staff':staff})
        else:
            messages.error(request,'No Staff Found')
            return redirect('overall_profile')
 

def get_student_profile(request):
    if request.method!='POST':
        return render(request, 'hod_template/get_profile.html',{'type':'student'})
    else:
        student_id = request.POST.get('student_id')
        print(student_id)
        student = Student.objects.filter(register_number = student_id).first() or Student.objects.filter(roll_number=student_id).first()
        if student:
            return render(request, 'student_template/student_profile.html',{'student':student})
        else:
            messages.error(request,'No Student Found')
            return redirect('overall_profile')


def admin_view_overall_attendance(request):
    user = get_object_or_404(Admin, user=request.user)
    form = OverallAttendanceSelectionForm(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            from_date = form.cleaned_data.get('from_date')
            print(from_date)
            to_date = form.cleaned_data.get('to_date')
            print(to_date)
            department = form.cleaned_data.get('department')
            print(department)
            class_id = form.cleaned_data.get('class_name')
            print(class_id)
            roll_number = form.get_roll_number()
            print(roll_number)
            
            # Validate date range
            if from_date and to_date and from_date > to_date:
                messages.error(request, "From date must be less than to date and they should not be equal.")
                return render(request, "hod_template/admin_view_overall_attendance.html", {'form': form, 'page_title': 'Select Attendance Details'})
            
            overall_attendance = {}

            if class_id:
                students = Student.objects.filter(class_name=class_id)
            elif department:
                students = Student.objects.filter(class_name__department=department)
            elif roll_number:
                students = Student.objects.filter(roll_number=roll_number)
            else:
                students = Student.objects.all()
            print(students)
            
            if from_date and to_date:
                current_date = from_date
                while current_date <= to_date:
                    attendance_array = {}
                    if students.exists():
                        print(students)
                        for student in students:
                            print(student)
                            status = Attendance.objects.filter(date=current_date, student=student).first()
                            if status:
                                attendance_array[student.roll_number] = {
                                    1: 'Present',
                                    2: 'OD Internal',
                                    3: 'OD External',
                                    0: 'Absent',
                                    4: 'Pending'
                                }.get(status.status, 'Unknown')
                    if attendance_array:
                        overall_attendance[str(current_date)] = attendance_array
                    current_date += datetime.timedelta(days=1)
            else:
                messages.error(request, "Please provide a valid date range.")
                return render(request, "hod_template/admin_view_overall_attendance.html", {'form': form, 'page_title': 'Select Attendance Details'})
            print(overall_attendance)
            dates = list(overall_attendance.keys())
            roll_numbers = list(overall_attendance[dates[0]].keys()) if dates else []
            data_to_send={
                'overall_attendance': overall_attendance,
                'class_name': str(class_id) if class_id else 'All Classes',
                'dates': dates,
                'roll_numbers': roll_numbers
            }
            print(data_to_send)
            return render(request, "hod_template/overall_attendance_view_page.html",data_to_send )
        else:
            messages.error(request, "Invalid Data Provided")
    
    context = {'form': form, 'page_title': 'Select Attendance Details'}
    return render(request, "hod_template/admin_view_overall_attendance.html", context)


def add_result(request):
    exam_type = ExamDetail.objects.annotate(result_count=Count('examresult')).filter(result_count=0)
    class_lists = ClassList.objects.all()
    if exam_type and class_lists:
        return render(request,'hod_template/add_result.html',{'exam_type':exam_type, 'class_lists':class_lists})
    messages.error(request, 'Exam or Class Not Found')
    return redirect('admin_home')


def upload_result(request):
    if request.method == 'POST':
        exam_type_id = request.POST.get('exam_type')
        class_id = request.POST.get('class_id')

        # Fetch the ExamDetail and ClassList instances
        exam_detail = get_object_or_404(ExamDetail, id=exam_type_id)
        class_name = get_object_or_404(ClassList, id=class_id)
        periods = Period.objects.filter(class_name=class_name)
        subjects = [period.subject for period in periods]
        expected_headers = ['roll_number'] + [subject.name for subject in subjects]

        # Check if a file was uploaded
        if 'result_file' not in request.FILES:
            messages.error(request, 'No file uploaded')
            return redirect('add_result')

        result_file = request.FILES['result_file']

        # Read the uploaded CSV file
        data = []
        try:
            csv_file = csv.reader(result_file.read().decode('utf-8').splitlines())
            headers = next(csv_file)
            
            # Validate headers
            if headers != expected_headers:
                messages.error(request, 'Invalid CSV structure')
                return redirect('add_result')

            for row in csv_file:
                data.append(row)
        except Exception as e:
            messages.error(request, f'Error reading CSV file: {e}')
            return redirect('add_result')

        # Validate the data rows
        for row in data:
            if len(row) != len(expected_headers):
                messages.error(request, 'Invalid row length in CSV file')
                return redirect('add_result')
            roll_number = row[0]
            marks = row[1:]
            # Validate roll number
            student = Student.objects.filter(roll_number=roll_number, class_name=class_name).first()
            if not student:
                messages.error(request, f'Invalid roll number: {roll_number}')
                return redirect('add_result')
            # Validate marks
            for mark in marks:
                if not mark.isdigit() or not (0 <= int(mark) <= 100):
                    messages.error(request, f'Invalid mark for roll number {roll_number}: {mark}')
                    return redirect('add_result')

        # Insert validated data into the ExamResult model
        for row in data:
            roll_number = row[0]
            student = Student.objects.get(roll_number=roll_number, class_name=class_name)
            marks_dict = {}
            for index, subject in enumerate(subjects):
                marks_dict[subject.name] = int(row[index + 1])
            
            # Create or update the ExamResult entry
            ExamResult.objects.update_or_create(
                exam_detail=exam_detail,
                student=student,
                defaults={'marks': marks_dict}
            )

        messages.success(request, 'Marks uploaded successfully')
        return redirect('add_result')
    else:
        exam_type = ExamDetail.objects.all()
        class_lists = ClassList.objects.all()
        if exam_type and class_lists:
            return render(request,'hod_template/add_result.html',{'exam_type':exam_type, 'class_lists':class_lists})
        else:
            return HttpResponse('Invalid request method', status=405)


def view_result(request):
    form = ResultViewForm(request.POST or None)
    student = get_object_or_404(Admin, user_id=request.user.id)
    context = {
        'form': form,
        'page_title': 'View Result'
    }
    
    if request.method == 'POST' and form.is_valid():
        try:
            exam_type = form.cleaned_data['exam_type']
            class_name = form.cleaned_data['class_name']        
            students = Student.objects.filter(class_name=class_name)
            marks_data = {}
            subjects = set()  # To collect all subject names
            
            for student in students:
                exam_result = ExamResult.objects.filter(exam_detail=exam_type, student=student).first()
                if exam_result and exam_result.marks:
                    marks_data[student.roll_number] = exam_result.marks
                    subjects.update(exam_result.marks.keys())
            
            subjects = sorted(subjects)  # Sort the subjects for consistent ordering
            
            context['marks_data'] = marks_data
            context['students'] = students
            context['subjects'] = subjects
            context['exam_type'] = exam_type
            context['class_name'] = class_name
            if len(marks_data)==0 :
                messages.error(request, 'No Data Found')
                return redirect('view_result')
            return render(request, "hod_template/view_result.html", context)
        except Exception as e:
            print(e)
            messages.error(request, "Could not retrieve results!")
    elif request.method == 'POST':
        messages.error(request, "Form has errors!")
        
    return render(request, "hod_template/view_result_form.html", context)


def download_result_template(request):
    if request.method == 'POST':
        exam_type_id = request.POST.get('exam_type')
        class_id = request.POST.get('class_id')

        # Fetch the ExamDetail and ClassList instances
        exam_detail = get_object_or_404(ExamDetail, id=exam_type_id)
        class_name = get_object_or_404(ClassList, id=class_id)

        periods = Period.objects.filter(class_name=class_name)
        print(periods)
        print(f"Class ID: {class_name}, Exam Type ID: {exam_detail}")
        # Validate that the class matches the exam's department and semester
        if class_name.department != exam_detail.department or class_name.semester != exam_detail.semester:
            messages.error(request, 'The Exam is Not For This Class')
            return redirect('upload_result')

        # Fetch the students in the specified class
        students = Student.objects.filter(class_name=class_name)

        # Prepare the HTTP response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{class_name.department.name}_{class_name.semester}-semester_result.csv"'

        # Create a CSV writer object
        writer = csv.writer(response)
        header = ['roll_number']
        for period in periods:
            header.append(period.subject.name)
        # Write the header row
        writer.writerow(header)

        # Write the student data rows
        for student in students:
            writer.writerow([str(student.roll_number), ''])

        return response
    else:
        messages.error(request, 'Invalid Request')
        return redirect('upload_result')
    

def admin_view_certificate(request):
    user = get_object_or_404(Admin, user=request.user)
    certificates = Certificate.objects.all()
    return render(request,'hod_template/admin_view_certificate.html',{'page_title': 'Student Certificates','certificates':certificates})


def admin_view_profile(request):
    admin = get_object_or_404(Admin, user=request.user)
    form = AdminForm(request.POST or None, request.FILES or None,
                     instance=admin)
    context = {'form': form,
               'page_title': 'View/Edit Profile'
               }
    if request.method == 'POST':
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get('first_name')
                last_name = form.cleaned_data.get('last_name')
                password = form.cleaned_data.get('password') or None
                passport = request.FILES.get('profile_pic') or None
                custom_user = admin.user
                if password != None:
                    custom_user.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    # passport_url = fs.url(filename)
                    custom_user.profile_pic = filename
                custom_user.first_name = first_name
                custom_user.last_name = last_name
                custom_user.save()
                messages.success(request, "Profile Updated!")
                return redirect(reverse('admin_view_profile'))
            else:
                messages.error(request, "Invalid Data Provided")
        except Exception as e:
            messages.error(
                request, "Error Occured While Updating Profile " + str(e))
    return render(request, "hod_template/admin_view_profile.html", context)


def admin_notify_staff(request):
    user = get_object_or_404(Admin, user=request.user)
    staff = CustomUser.objects.filter(user_type=2)
    context = {
        'page_title': "Send Notifications To Staff",
        'allStaff': staff
    }
    return render(request, "hod_template/staff_notification.html", context)


def admin_notify_student(request):
    user = get_object_or_404(Admin, user=request.user)
    student = CustomUser.objects.filter(user_type=3)
    context = {
        'page_title': "Send Notifications To Students",
        'students': student
    }
    return render(request, "hod_template/student_notification.html", context)


def exam_filter_page(request):
    form = ExamDetailForm()
    context={
        'form':form,
        'page_title': 'Question Paper'
    }
    if request.method == 'POST':
        form = ExamDetailForm(request.POST)
        if form.is_valid():
            department = form.cleaned_data['department']
            subject = form.cleaned_data['subject']
            semester = form.cleaned_data['semester']
            exam_type = form.cleaned_data['exam_type']
            exam_detail = ExamDetail.objects.filter(department=department, semester=semester, exam_type=exam_type).first()
            q_paper = Question.objects.filter(exam_detail=exam_detail, subject=subject).first()

            if q_paper:
                return render(request, "hod_template/question_paper_template.html", {'q_paper':q_paper, 'exam': q_paper.exam_detail})
            else:
                messages.error(request, 'No Question Paper Found.')
                
        else:
            messages.error(request, 'There was an error with your form submission.')
    
    return render(request, 'hod_template/exam_filter_page.html', {'form': form,'page_title':'Question Paper Details'})


def hod_class_timetable(request):
    if request.method!='POST':
        department=Department.objects.all()
        class_names = ClassList.objects.all()
        context={
            'department':department,
            'class_names':class_names,
            'type':'class'
        }
        return render(request, 'hod_template/get_timetable_form.html',context)
    else:
        class_id = request.POST.get('class_id')
        print(class_id)
        timetable = TimeTable.objects.filter(class_name=class_id).first()
        if not timetable:
            messages.error(request,'No Timetable found for the class')
            return redirect('overall_timetable')

        return render(request, 'student_template/student_timetable.html', {'timetable':timetable, 'class_name':str(timetable.class_name)})


def hod_staff_timetable(request):
    user = get_object_or_404(Admin, user=request.user)
    if request.method!='POST':
        return render(request, 'hod_template/get_timetable_form.html', {'type':'staff'})
    else:
        faculty_id = request.POST.get('faculty_id')
        staff = Staff.objects.filter(faculty_id=faculty_id).first()

        if not staff:
            messages.error(request, "No Staff Found")
            return redirect('overall_timetable')

        periods = Period.objects.filter(staff=staff)
        staff_timetable = {}
        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
        period_count = ['1','2','3','4','5','6','7','8']

        for day in days:
            for count in period_count:
                column = f"{day}_{count}"
                print(column)

                timetable_list = TimeTable.objects.filter(
                    department=staff.department,
                    class_name__in=[period.class_name for period in periods]
                )
                found = False
                for timetable in timetable_list:
                    print(getattr(timetable,column).staff, '\n', type(getattr(timetable,column).staff))
                    print(staff, '\n', type(staff))
                    if getattr(timetable,column).staff == staff:
                        staff_timetable[column] = timetable.class_name
                        found = True
                        break
                if found == False:
                    staff_timetable[column] = 'Free Period'
        print(staff_timetable)

        if not timetable:
            return render(request, 'error.html', {'message': 'Timetable not found for this staff member'})
            

        return render(request, 'staff_template/staff_timetable.html', {'staff_timetable': staff_timetable, 'staff_name': staff.user})


@csrf_exempt
def student_feedback_message(request):
    user = get_object_or_404(Admin, user=request.user)
    if request.method != 'POST':
        feedbacks = FeedbackStudent.objects.all()
        context = {
            'feedbacks': feedbacks,
            'page_title': 'Student Feedback Messages'
        }
        return render(request, 'hod_template/student_feedback_template.html', context)
    else:
        feedback_id = request.POST.get('id')
        try:
            feedback = get_object_or_404(FeedbackStudent, id=feedback_id)
            reply = request.POST.get('reply')
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def staff_feedback_message(request):
    user = get_object_or_404(Admin, user=request.user)
    if request.method != 'POST':
        feedbacks = FeedbackStaff.objects.all()
        context = {
            'feedbacks': feedbacks,
            'page_title': 'Staff Feedback Messages'
        }
        return render(request, 'hod_template/staff_feedback_template.html', context)
    else:
        feedback_id = request.POST.get('id')
        try:
            feedback = get_object_or_404(FeedbackStaff, id=feedback_id)
            reply = request.POST.get('reply')
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def view_staff_leave(request):
    user = get_object_or_404(Admin, user=request.user)
    if request.method != 'POST':
        allLeave = LeaveReportStaff.objects.all()
        context = {
            'allLeave': allLeave,
            'page_title': 'Leave Applications From Staff'
        }
        return render(request, "hod_template/staff_leave_view.html", context)
    else:
        id = request.POST.get('id')
        status = request.POST.get('status')
        if (status == '1'):
            status = 1
            rejection_reason=None
        else:
            status = -1
            rejection_reason=request.POST.get('reason','')
        try:
            leave = get_object_or_404(LeaveReportStaff, id=id)
            leave.status = status
            leave.rejection_reason=rejection_reason
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


@csrf_exempt
def view_student_leave(request):
    user = get_object_or_404(Admin, user=request.user)
    if request.method != 'POST':
        allLeave = LeaveReportStudent.objects.all()
        context = {
            'allLeave': allLeave,
            'page_title': 'Leave Applications From Students'
        }
        return render(request, "hod_template/student_leave_view.html", context)
    else:
        id = request.POST.get('id')
        status = request.POST.get('status')
        if (status == '1'):
            status = 1
            rejection_reason=None
        else:
            status = -1
            rejection_reason=request.POST.get('reason','')
        try:
            leave = get_object_or_404(LeaveReportStudent, id=id)
            leave.status = status
            leave.rejection_reason=rejection_reason
            print(leave)
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


@csrf_exempt
def send_student_notification(request):
    user = get_object_or_404(Admin, user=request.user)
    id = request.POST.get('id')
    message = request.POST.get('message')
    student = get_object_or_404(Student, user_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            'notification': {
                'title': "Student Management System",
                'body': message,
                'click_action': reverse('student_view_notification'),
                'icon': static('dist/img/AdminLTELogo.png')
            },
            'to': student.user.fcm_token
        }
        headers = {'Authorization':
                   'key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB',
                   'Content-Type': 'application/json'}
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStudent(student=student, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


@csrf_exempt
def send_staff_notification(request):
    user = get_object_or_404(Admin, user=request.user)
    id = request.POST.get('id')
    message = request.POST.get('message')
    staff = get_object_or_404(Staff, user_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            'notification': {
                'title': "Student Management System",
                'body': message,
                'click_action': reverse('staff_view_notification'),
                'icon': static('dist/img/AdminLTELogo.png')
            },
            'to': staff.user.fcm_token
        }
        headers = {'Authorization':
                   'key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB',
                   'Content-Type': 'application/json'}
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStaff(staff=staff, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def download_student_template(request):
    user = get_object_or_404(Admin, user=request.user)
    # Create the HTTP response object with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_template.csv"'

    # Create a CSV writer object
    writer = csv.writer(response)

    # Write the header row
    writer.writerow([
        'email',
        'password',
        'gender',
        'address',
        'department',
        'class',
        'register_number',
        'roll_number'
    ])

    return response


def download_staff_template(request):
    user = get_object_or_404(Admin, user=request.user)
    # Create the HTTP response object with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_template.csv"'

    # Create a CSV writer object
    writer = csv.writer(response)

    # Write the header row
    writer.writerow([
        'email',
        'password',
        'gender',
        'address',
        'department'
    ])

    return response


@require_GET
def download_single_day_attendance(request):
    user = get_object_or_404(Admin, user=request.user)
    attendance_list_str = request.GET.get('attendance_list')
    if attendance_list_str:
        try:
            attendance_list = json.loads(attendance_list_str)
        except json.JSONDecodeError:
            attendance_list = []
    else:
        attendance_list = []
    roll_num_of_one_student = attendance_list[0][0]
    student = get_object_or_404(Student, roll_number=roll_num_of_one_student)

    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = str(student.class_name)+" - Attendance"

    # Define fill colors
    # Orange background for Absent
    absent_fill = PatternFill(start_color="FF4400",
                              end_color="FF4400", fill_type="solid")
    # Yellow background for OD Internal and External
    od_internal_external_fill = PatternFill(
        start_color="FFCC33", end_color="FFCC33", fill_type="solid")
    pending_fill = PatternFill(
        start_color="00555E", end_color="00555E", fill_type="solid")
    text_color_white = openpyxl.styles.Font(color="FFFFFF")  # White text color
    text_color_black = openpyxl.styles.Font(color="000000")  # Black text color

    # Write the header row
    headers = [
        'Roll Number',
        'Period 1',
        'Period 2',
        'Period 3',
        'Period 4',
        'Period 5',
        'Period 6',
        'Period 7',
        'Period 8'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        # Yellow background for headers
        # cell.fill = PatternFill(start_color="FFFF00",
        #                         end_color="FFFF00", fill_type="solid")

    # Write the data rows with conditional formatting
    for row_num, atten in enumerate(attendance_list, 2):
        for col_num, value in enumerate(atten, 1):
            cell = worksheet.cell(
                row=row_num, column=col_num, value=check_atten_with_number(value))
            if value == 0:
                cell.fill = absent_fill
                cell.font = text_color_white
            elif value in [2, 3]:
                cell.fill = od_internal_external_fill
                cell.font = text_color_black
            elif value == 4:
                cell.fill = pending_fill
                cell.font = text_color_white

    # Create the HTTP response object with Excel content type
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{str(student.class_name)} -attendance.xlsx"'

    # Save the workbook to the response object
    workbook.save(response)

    return response


def check_atten_with_number(atten):
    if atten != 0 and atten != 1 and atten != 2 and atten != 3 and atten != 4:
        return atten
    elif atten == 0:
        return 'Absent'
    elif atten == 1:
        return 'Present'
    elif atten == 2:
        return 'On Duty Internal'
    elif atten == 3:
        return 'On Duty External'
    elif atten == 4:
        return 'Pending'


@require_GET
def download_overall_day_attendance(request):
    print('frbtdg\n\n\n')
    user = get_object_or_404(Admin, user=request.user)
    overall_attendance = json.loads(request.GET.get('overall_attendance', '{}'))

    # Add headers
    dates = list(overall_attendance.keys())
    roll_numbers = list(next(iter(overall_attendance.values())).keys())
    student = get_object_or_404(Student, roll_number=roll_numbers[0])
    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Attendance of {student.class_name}'


    header = ['Roll Number'] + dates
    sheet.append(header)

    # Add data
    for reg in roll_numbers:
        row = [reg]
        for date in dates:
            row.append(overall_attendance[date].get(reg, ''))
        sheet.append(row)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={student.class_name}_overall_attendance.xlsx'

    # Save workbook to response
    workbook.save(response)
    return response